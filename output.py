
###=======BUILT FROM SCRATCH BY MACAULAY OLADIMEJI SAHEED / +2348160927317=====================##################


import tkinter as tk
from tkinter import filedialog
import openpyxl
# from tkcalendar import DateEntry
from tkinter import ttk
from tkinter import messagebox
from openpyxl.formula.translate import Translator
from tkinter import messagebox, simpledialog
from tkhtmlview import HTMLLabel
import subprocess
from PIL import Image, ImageDraw, ImageFont
import os

# import win32print  # Required for Windows systems



def load_last_number():
    if os.path.exists("./images/delete.txt"):
        with open("./images/delete.txt", "r") as file:
            return int(file.read())
    else:
        return 0
last_number = load_last_number()











# Function to generate the receipt
def generate_receipt():
    # Collect input values
    date = input_field.get()
    customer_name = input_field2.get()
    invoice_number = input_field1.get()
    total_amount = input_field55.get()

    # Validate input fields
    if not customer_name:
        messagebox.showinfo("Error", "Input Customer Name")
        return
    if not invoice_number:
        messagebox.showinfo("Error", "Input Invoice Number!")
        return
    if not total_amount:
        messagebox.showinfo("Error", "Calculate Total Amount!")
        return


    # # Display the receipt
    receipt_window = tk.Tk()
    receipt_window.geometry("300x600")
    receipt_window.title("Lacova Receipt")
    receipt_text = tk.Text(receipt_window, height=30, width=40)
    receipt_text.grid()
    receipt = f'       LACOVA INFINITY     \n'
    receipt += f'      GLOBAL ENTERPRISES     \n'
    receipt += f' Omoleye Area, Mechanic Village,\n Fodacis, Adeoyo, Oluyole Ibadan\n'
    receipt += f'Tel: 08053625844, 09018199769     \n\n'
    
    receipt += f"Customer Name:      {customer_name}\n"
    receipt += f"Invoice Number:    {invoice_number}\n"
    receipt += f"Date:      {date}\n\n"

    receipt += "QTY  ITEM    PRICE\n\n"

    if input_field4.get() and input_field3.get():
        receipt += f"{input_field3.get()}   3.4Kg    #{float(input_field4.get()) * float(input_field3.get())}\n"
    if input_field6.get() and input_field5.get():
        receipt += f"{input_field5.get()}   3.5Kg    #{float(input_field6.get()) * float(input_field5.get())}\n"
    if input_field8.get() and input_field7.get():
        receipt += f"{input_field7.get()}   3.6Kg    #{float(input_field8.get()) * float(input_field7.get())}\n"
    if input_field10.get() and input_field9.get():
        receipt += f"{input_field9.get()}   3.7Kg    #{float(input_field10.get()) * float(input_field9.get())}\n"
    if input_field12.get() and input_field11.get():
        receipt += f"{input_field11.get()}   3.8Kg    #{float(input_field12.get()) * float(input_field11.get())}\n"
    if input_field14.get() and input_field13.get():
        receipt += f"{input_field13.get()}   3.9Kg    #{float(input_field14.get()) * float(input_field13.get())}\n"
    if input_field16.get() and input_field15.get():
        receipt += f"{input_field15.get()}   4.0Kg    #{float(input_field16.get()) * float(input_field15.get())}\n"
    if input_field18.get() and input_field17.get():
        receipt += f"{input_field17.get()}   4.1Kg    #{float(input_field18.get()) * float(input_field17.get())}\n"
    if input_field20.get() and input_field19.get():
        receipt += f"{input_field19.get()}   4.2Kg    #{float(input_field20.get()) * float(input_field19.get())}\n"
    if input_field22.get() and input_field21.get():
        receipt += f"{input_field21.get()}   4.3Kg    #{float(input_field22.get()) * float(input_field21.get())}\n"
    if input_field24.get() and input_field23.get():
        receipt += f"{input_field23.get()}   4.4Kg    #{float(input_field24.get()) * float(input_field23.get())}\n"
    if input_field26.get() and input_field25.get():
        receipt += f"{input_field25.get()}   4.5Kg    #{float(input_field26.get()) * float(input_field25.get())}\n"
    if input_field28.get() and input_field27.get():
        receipt += f"{input_field27.get()}   4.6Kg    #{float(input_field28.get()) * float(input_field27.get())}\n"
    if input_field30.get() and input_field29.get():
        receipt += f"{input_field29.get()}   4.7Kg    #{float(input_field30.get()) * float(input_field29.get())}\n"
    if input_field32.get() and input_field31.get():
        receipt += f"{input_field31.get()}   4.8Kg    #{float(input_field32.get()) * float(input_field31.get())}\n"
    if input_field34.get() and input_field33.get():
        receipt += f"{input_field33.get()}   4.9Kg    #{float(input_field34.get()) * float(input_field33.get())}\n"
    if input_field36.get() and input_field35.get():
        receipt += f"{input_field35.get()}   5.0Kg    #{float(input_field36.get()) * float(input_field35.get())}\n"
    if input_field38.get() and input_field37.get():
        receipt += f"{input_field37.get()}   5.2Kg    #{float(input_field38.get()) * float(input_field37.get())}\n"
    if input_field40.get() and input_field39.get():
        receipt += f"{input_field39.get()}   5.5Kg    #{float(input_field40.get()) * float(input_field39.get())}\n"
    if input_field42.get() and input_field41.get():
        receipt += f"{input_field41.get()}   5.8Kg    #{float(input_field42.get()) * float(input_field41.get())}\n"
    if input_field46.get() and input_field45.get():
        receipt += f"{input_field45.get()}   6.0Kg    #{float(input_field46.get()) * float(input_field45.get())}\n"
    if input_field48.get() and input_field47.get():
        receipt += f"{input_field47.get()}   Oloja    #{float(input_field48.get()) * float(input_field47.get())}\n"
    if input_field50.get() and input_field49.get():
        receipt += f"{input_field49.get()}   Ice      #{float(input_field50.get()) * float(input_field49.get())}\n"
    if input_field52.get() and input_field51.get():
        receipt += f"{input_field51.get()}   Dispenser#{float(input_field52.get()) * float(input_field51.get())}\n"
    if input_field54.get() and input_field53.get():
        receipt += f"{input_field53.get()}   Small    #{float(input_field54.get()) * float(input_field53.get())}\n\n"
    receipt += f"  _______________________________________\n"
    receipt += f"  Total Amount:     #{total_amount}\n"
    receipt += f"  _______________________________________\n\n"
    receipt += f"  Amount Paid:      #{float(input_field61.get())}\n"
    balance = float(input_field61.get()) - float(input_field55.get())
    if balance < 0: 
        receipt += f"  Outstanding:      #{float(input_field63.get())}\n\n"
    else:
        receipt += f"  Prepayment:      #{float(input_field64.get())}\n\n"

    newbalances = float(input_field65.get())
    if newbalances < 0:
        receipt += f"  Total Outstanding: #{float(input_field65.get())}\n\n"
    else:
        receipt += f"  Total Prepayment: #{float(input_field65.get())}\n\n"



    receipt += f'  Sales Attendant:    \n\n\n'
    receipt += f'     Thanks for your patronage'

    # Display the receipt
    receipt_text.delete(1.0, tk.END)  # Clear previous receipt
    receipt_text.insert(tk.END, receipt)


        # Enable the print button



    # Enable the print button
    print_button = tk.Button(receipt_window, text="Print Receipt", command=print_receipt, state="disabled", bg="green", fg="white", font=("Helvetica", 12), width=10, bd=0)
    print_button.grid()
    print_button.configure(state="normal")
    # print_button.configure(state="normal")
    print_button["command"] = lambda: print_receipt(image_filename)

    # Create an image of the receipt
    image_width = 200
    image_height = 500
    receipt_image = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(receipt_image)
    font = ImageFont.load_default()

    # Draw the receipt content on the image
    y = 10
    for line in receipt.split("\n"):
        draw.text((10, y), line, fill="black", font=font)
        y += 15

    # Save the receipt image
    image_filename = f"receipts/receipt_{customer_name}_{invoice_number}.png"
    receipt_image.save(image_filename)
    download_button = tk.Button(receipt_window, text="Download Receipt", command=download_receipt, state="disabled", bg="blue", fg="white", font=("Helvetica", 12), width=15, bd=0)
    download_button.grid()
    download_button.configure(state="normal")
    # Enable the download button
    # download_button.configure(state="normal")

def download_receipt():
    # Open the receipt image file
    invoice_number = input_field1.get()
    image_filename = "./receipts/receipt{invoice_number}.png"
    image = Image.open(image_filename)

    # Prompt the user to save the image file
    save_filename = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG Image", "*.png")])
    if save_filename:
        image.save(save_filename)

# Function to generate the waybill 
def generate_Waybill():
    # Collect input values
    date = input_field.get()
    customer_name = input_field2.get()
    invoice_number = input_field1.get()
    total_amount = input_field55.get()

    # Validate input fields
    if not customer_name:
        messagebox.showinfo("Error", "Input Customer Name")
        return
    if not invoice_number:
        messagebox.showinfo("Error", "Input Invoice Number!")
        return
    if not total_amount:
        messagebox.showinfo("Error", "Calculate Total Amount!")
        return


    # # Display the waybill
    Waybill_window = tk.Tk()
    Waybill_window.geometry("300x600")
    Waybill_window.title("Lacova Waybill")
    Waybill_text = tk.Text(Waybill_window, height=30, width=40)
    Waybill_text.grid()
    Waybill = f'      LACOVA INFINITY     \n'
    Waybill += f'    GLOBAL ENTERPRISES     \n'
    Waybill += f' Omoleye Area, Mechanic Village,\n Fodacis, Adeoyo, Oluyole Ibadan\n'
    Waybill += f'Tel: 0805 362 5844, 0901 819 9769     \n\n'
    
    Waybill += f"Customer Name:      {customer_name}\n"
    Waybill += f"Waybill Number:    {invoice_number}\n"
    Waybill += f"Date:      {date}\n\n"

    Waybill += "QTY         ITEM        \n\n"

    if input_field4.get() and input_field3.get():
        Waybill += f"  {input_field3.get()}      3.4Kg        \n"
    if input_field6.get() and input_field5.get():
        Waybill += f"  {input_field5.get()}      3.5Kg        \n"
    if input_field8.get() and input_field7.get():
        Waybill += f"  {input_field7.get()}      3.6Kg        \n"
    if input_field10.get() and input_field9.get():
        Waybill += f"  {input_field9.get()}      3.7Kg         \n"
    if input_field12.get() and input_field11.get():
        Waybill += f"  {input_field11.get()}      3.8Kg        \n"
    if input_field14.get() and input_field13.get():
        Waybill += f"  {input_field13.get()}      3.9Kg        \n"
    if input_field16.get() and input_field15.get():
        Waybill += f"  {input_field15.get()}      4.0Kg        \n"
    if input_field18.get() and input_field17.get():
        Waybill += f"  {input_field17.get()}      4.1Kg        \n"
    if input_field20.get() and input_field19.get():
        Waybill += f"  {input_field19.get()}      4.2Kg        \n"
    if input_field22.get() and input_field21.get():
        Waybill += f"  {input_field21.get()}      4.3Kg        \n"
    if input_field24.get() and input_field23.get():
        Waybill += f"  {input_field23.get()}      4.4Kg        \n"
    if input_field26.get() and input_field25.get():
        Waybill += f"  {input_field25.get()}      4.5Kg        \n"
    if input_field28.get() and input_field27.get():
        Waybill += f"  {input_field27.get()}      4.6Kg        \n"
    if input_field30.get() and input_field29.get():
        Waybill += f"  {input_field29.get()}      4.7Kg        \n"
    if input_field32.get() and input_field31.get():
        Waybill += f"  {input_field31.get()}      4.8Kg        \n"
    if input_field34.get() and input_field33.get():
        Waybill += f"  {input_field33.get()}      4.9Kg        \n"
    if input_field36.get() and input_field35.get():
        Waybill += f"  {input_field35.get()}      5.0Kg        \n"
    if input_field38.get() and input_field37.get():
        Waybill += f"  {input_field37.get()}      5.2Kg        \n"
    if input_field40.get() and input_field39.get():
        Waybill += f"  {input_field39.get()}      5.5Kg        \n"
    if input_field42.get() and input_field41.get():
        Waybill += f"  {input_field41.get()}      5.8Kg        \n"
    if input_field46.get() and input_field45.get():
        Waybill += f"  {input_field45.get()}      6.0Kg        \n"
    if input_field48.get() and input_field47.get():
        Waybill += f"  {input_field47.get()}      Oloja        \n"
    if input_field50.get() and input_field49.get():
        Waybill += f"  {input_field49.get()}      Ice          \n"
    if input_field52.get() and input_field51.get():
        Waybill += f"  {input_field51.get()}      Dispenser    \n"
    if input_field54.get() and input_field53.get():
        Waybill += f"  {input_field53.get()}      Small        \n\n"

    Waybill += f"_________________________________\n\n\n"
    
    Waybill += f'Sales Attendant:    \n\n\n'
    Waybill += f'   Thanks for your patronage'

    # Display the Waybill
    Waybill_text.delete(1.0, tk.END)  # Clear previous Waybill
    Waybill_text.insert(tk.END, Waybill)

    # Enable the print button
    print_button = tk.Button(Waybill_window, text="Print Waybill", command=print_Waybill, state="disabled", bg="green", fg="white", font=("Helvetica", 12), width=10, bd=0)
    print_button.grid()
    print_button.configure(state="normal")


def print_receipt(image_filename):
    # Print the receipt using the lp command
    try:
        subprocess.run(["lp", image_filename])
        messagebox.showinfo("Print Receipt", "Receipt sent to printer successfully!")
    except FileNotFoundError:
        messagebox.showerror("Print Receipt", "Printer command (lp) not found!")

def print_Waybill():
    # Get the receipt HTML
    receipt_label = tk.StringVar()
    receipt_html = receipt_label.get()

    # Get the printer name from the user
    printer_name = simpledialog.askstring("Printer Selection", "Enter Printer Name:")

    try:
        # Print the receipt using the 'lpr' command-line utility
        lpr_process = subprocess.Popen(["lpr", "-P", printer_name], stdin=subprocess.PIPE)
        lpr_process.communicate(Waybill.encode())

        messagebox.showinfo("Print", "Receipt sent to the printer.")

    except Exception as e:
        messagebox.showinfo("Error", str(e))

def check_credentials():
    # Check if the entered username and password are correct
    if username_entry.get() == "lacova" and password_entry.get() == "lacova":
        # If they are correct, hide the login window and show the main application window
        password_entry.delete(0, tk.END)
        login_window.withdraw()
        window.deiconify()
    else:
        # If they are incorrect, show an error message as a pop-up
        messagebox.showerror("Error", "Incorrect username or password")
        # Clear the input fields
        username_entry.delete(0, tk.END)
        password_entry.delete(0, tk.END)  
def logout():
    # Hide the main application window and show the login window again
    window.withdraw()
    login_window.deiconify()
# Create the main application window
def calculate_total():
    # Initialize the total variable to 0
    total = 0
    # Calculate the total for each input field that contains a value
    if input_field4.get() and input_field3.get():
        total += float(input_field4.get()) * float(input_field3.get())
    if input_field6.get() and input_field5.get():
        total += float(input_field6.get()) * float(input_field5.get())
    if input_field8.get() and input_field7.get():
        total += float(input_field8.get()) * float(input_field7.get())
    if input_field10.get() and input_field9.get():
        total += float(input_field10.get()) * float(input_field9.get())
    if input_field12.get() and input_field11.get():
        total += float(input_field12.get()) * float(input_field11.get())
    if input_field14.get() and input_field13.get():
        total += float(input_field14.get()) * float(input_field13.get())
    if input_field16.get() and input_field15.get():
        total += float(input_field16.get()) * float(input_field15.get())
    if input_field18.get() and input_field17.get():
        total += float(input_field18.get()) * float(input_field17.get())
    if input_field20.get() and input_field19.get():
        total += float(input_field20.get()) * float(input_field19.get())
    if input_field22.get() and input_field21.get():
        total += float(input_field22.get()) * float(input_field21.get())
    if input_field24.get() and input_field23.get():
        total += float(input_field24.get()) * float(input_field23.get())
    if input_field26.get() and input_field25.get():
        total += float(input_field26.get()) * float(input_field25.get())
    if input_field28.get() and input_field27.get():
        total += float(input_field28.get()) * float(input_field27.get())
    if input_field30.get() and input_field29.get():
        total += float(input_field30.get()) * float(input_field29.get())
    if input_field32.get() and input_field31.get():
        total += float(input_field32.get()) * float(input_field31.get())
    if input_field34.get() and input_field33.get():
        total += float(input_field34.get()) * float(input_field33.get())
    if input_field36.get() and input_field35.get():
        total += float(input_field36.get()) * float(input_field35.get())
    if input_field38.get() and input_field37.get():
        total += float(input_field38.get()) * float(input_field37.get())
    if input_field40.get() and input_field39.get():
        total += float(input_field40.get()) * float(input_field39.get())
    if input_field42.get() and input_field41.get():
        total += float(input_field42.get()) * float(input_field41.get())
    # if input_field44.get() and input_field43.get():
    #     total += float(input_field44.get()) * float(input_field43.get())
    if input_field46.get() and input_field45.get():
        total += float(input_field46.get()) * float(input_field45.get())
    if input_field48.get() and input_field47.get():
        total += float(input_field48.get()) * float(input_field47.get())
    if input_field50.get() and input_field49.get():
        total += float(input_field50.get()) * float(input_field49.get())
    if input_field52.get() and input_field51.get():
        total += float(input_field52.get()) * float(input_field51.get())
    if input_field54.get() and input_field53.get():
        total += float(input_field54.get()) * float(input_field53.get())
    input_field55.delete(0, tk.END)
    input_field55.insert(0, total)
def calculate_total_invoice_amount():
    # Initialize the total variable to 0
    total_invoice = 0
    if input_field56.get():
        total_invoice += float(input_field56.get())
    if input_field57.get():
        total_invoice += float(input_field57.get())
    if input_field58.get():
        total_invoice += float(input_field58.get())
    if input_field59.get():
        total_invoice += float(input_field59.get())
    input_field61.delete(0, tk.END)
    input_field61.insert(0, total_invoice)
def calculate_balances():
    balance = float(input_field61.get()) - float(input_field55.get())
    if balance < 0:
        #input in outstanding field
        input_field63.delete(0, tk.END)
        input_field64.delete(0, tk.END)
        input_field65.delete(0, tk.END)
        input_field63.insert(0, balance)
        newbalance = float(input_field62.get()) + float(input_field63.get())
        input_field65.insert(0, newbalance)
    else:
        #input in prepayment field
        input_field63.delete(0, tk.END)
        input_field64.delete(0, tk.END)
        input_field65.delete(0, tk.END)
        input_field64.insert(0, balance)
        newbalance = float(input_field62.get()) + float(input_field64.get())
        input_field65.insert(0, newbalance)
def save_data():
    if not input_field2.get():
        not_submitted_label.config(text="Fill the neccesary fields")
        messagebox.showinfo("Error", "Input Customer Name")
        # window.after(3000, lambda: messagebox._show("",""))
        not_submitted_label.after(3000, lambda: not_submitted_label.config(text="")) # Remove the text after 5 seconds
        return  # Do nothing if input field is empty
    if not input_field1.get():
        not_submitted_label.config(text="Fill the neccesary fields")
        messagebox.showinfo("Error", "Input invoice Number!")
        not_submitted_label.after(3000, lambda: not_submitted_label.config(text="")) # Remove the text after 5 seconds
        return  # Do nothing if input field is empty
    if not input_field55.get():
        not_submitted_label.config(text="Fill the neccesary fields")
        messagebox.showinfo("Error", "Calculate Total Goods Price!")
        not_submitted_label.after(3000, lambda: not_submitted_label.config(text="")) # Remove the text after 5 seconds
        return  # Do nothing if input field is empty
    if not input_field61.get():
        not_submitted_label.config(text="Fill the neccesary fields")
        messagebox.showinfo("Error", "Calculate Amount Paid!")
        not_submitted_label.after(3000, lambda: not_submitted_label.config(text=""))
        return
    # Load the existing workbook
    workbook = openpyxl.load_workbook('LACOVA_MANAGEMENT_ACCOUNT_2023.xlsx')
    # Select the sheet you want to modify
    sheet = workbook['Sales Invoice Listing']
    sheet1 = workbook['2023 TURNOVER']
    sheet2 = workbook['Debtors and Receivables']
    sheet3 = workbook['Cash Account Listing']
    sheet4 = workbook['CASH ACCOUNT']
    # Find the last row in the sheet
    max_row = sheet.max_row
    max_row1 = sheet1.max_row
    max_row2 = sheet2.max_row
    max_row3 = sheet3.max_row
    max_row4 = sheet4.max_row
    ###############Sales invoice listing################################################
    sheet.cell(row=max_row+1, column=1, value=input_field.get())
    sheet.cell(row=max_row+1, column=2, value=input_field1.get())
    sheet.cell(row=max_row+1, column=3, value=input_field2.get())
    sheet.cell(row=max_row+1, column=4, value=input_field3.get())    
    sheet.cell(row=max_row+1, column=5, value=input_field4.get())
    sheet.cell(row=max_row+1, column=6, value=input_field5.get())
    sheet.cell(row=max_row+1, column=7, value=input_field6.get())
    sheet.cell(row=max_row+1, column=8, value=input_field7.get())
    sheet.cell(row=max_row+1, column=9, value=input_field8.get())
    sheet.cell(row=max_row+1, column=10, value=input_field9.get())
    sheet.cell(row=max_row+1, column=11, value=input_field10.get())
    sheet.cell(row=max_row+1, column=12, value=input_field11.get())
    sheet.cell(row=max_row+1, column=13, value=input_field12.get())
    sheet.cell(row=max_row+1, column=14, value=input_field13.get())
    sheet.cell(row=max_row+1, column=15, value=input_field14.get())
    sheet.cell(row=max_row+1, column=16, value=input_field15.get())
    sheet.cell(row=max_row+1, column=17, value=input_field16.get())
    sheet.cell(row=max_row+1, column=18, value=input_field17.get())
    sheet.cell(row=max_row+1, column=19, value=input_field18.get())
    sheet.cell(row=max_row+1, column=20, value=input_field19.get())
    sheet.cell(row=max_row+1, column=21, value=input_field20.get())
    sheet.cell(row=max_row+1, column=22, value=input_field21.get())
    sheet.cell(row=max_row+1, column=23, value=input_field22.get())
    sheet.cell(row=max_row+1, column=24, value=input_field23.get())
    sheet.cell(row=max_row+1, column=25, value=input_field24.get())
    sheet.cell(row=max_row+1, column=26, value=input_field25.get())
    sheet.cell(row=max_row+1, column=27, value=input_field26.get())
    sheet.cell(row=max_row+1, column=28, value=input_field27.get())
    sheet.cell(row=max_row+1, column=29, value=input_field28.get())
    sheet.cell(row=max_row+1, column=30, value=input_field29.get())
    sheet.cell(row=max_row+1, column=31, value=input_field30.get())
    sheet.cell(row=max_row+1, column=32, value=input_field31.get())
    sheet.cell(row=max_row+1, column=33, value=input_field32.get())
    sheet.cell(row=max_row+1, column=34, value=input_field33.get())
    sheet.cell(row=max_row+1, column=35, value=input_field34.get())
    sheet.cell(row=max_row+1, column=36, value=input_field35.get())
    sheet.cell(row=max_row+1, column=37, value=input_field36.get())
    sheet.cell(row=max_row+1, column=38, value=input_field37.get())
    sheet.cell(row=max_row+1, column=39, value=input_field38.get())
    sheet.cell(row=max_row+1, column=40, value=input_field39.get())
    sheet.cell(row=max_row+1, column=41, value=input_field40.get())
    sheet.cell(row=max_row+1, column=42, value=input_field41.get())
    sheet.cell(row=max_row+1, column=43, value=input_field42.get())
    sheet.cell(row=max_row+1, column=44, value=input_field45.get())
    sheet.cell(row=max_row+1, column=45, value=input_field46.get())
    sheet.cell(row=max_row+1, column=46, value=input_field47.get())
    sheet.cell(row=max_row+1, column=47, value=input_field48.get())
    sheet.cell(row=max_row+1, column=48, value=input_field49.get())
    sheet.cell(row=max_row+1, column=49, value=input_field50.get())
    sheet.cell(row=max_row+1, column=50, value=input_field51.get())
    sheet.cell(row=max_row+1, column=51, value=input_field52.get())
    sheet.cell(row=max_row+1, column=52, value=input_field53.get())
    sheet.cell(row=max_row+1, column=53, value=input_field54.get())
    sheet.cell(row=max_row+1, column=54, value=input_field55.get())
    ######################################cash account listing########################################################################
    #Date
    sheet3.cell(row=max_row3+1, column=1, value=input_field.get())
    #customer name
    sheet3.cell(row=max_row3+1, column=2, value=input_field2.get())
    #details
    sheet3.cell(row=max_row3+1, column=3, value=input_field60.get())
    #invoice number
    sheet3.cell(row=max_row3+1, column=4, value=input_field1.get())  
    #amount on invoice
    sheet3.cell(row=max_row3+1, column=5, value=input_field55.get())
    #amount paid to fcmb
    sheet3.cell(row=max_row3+1, column=6, value=input_field56.get())
    #amount paid to polaris
    sheet3.cell(row=max_row3+1, column=7, value=input_field57.get())
    #amount paid to access
    sheet3.cell(row=max_row3+1, column=8, value=input_field58.get())
    #amount paid in cash
    sheet3.cell(row=max_row3+1, column=9, value=input_field59.get())
    #amount paid in total
    sheet3.cell(row=max_row3+1, column=10, value=input_field61.get())
    #amount outstanding
    sheet3.cell(row=max_row3+1, column=11, value=input_field63.get())
    #amount prepayment
    sheet3.cell(row=max_row3+1, column=12, value=input_field64.get())
    #amount new balance
    # sheet3.cell(row=max_row3+1, column=13, value=input_field65.get())
    old_balancess = float(input_field62.get())
    formula_newbalance = f'={old_balancess}+K{max_row3+1}+L{max_row3+1}'
    sheet3[f"M{max_row3+1}"] = formula_newbalance
    ######################################Cash Account########################################################################
    #date
    sheet4.cell(row=max_row4+1, column=1, value=input_field.get())
    #invoice number
    sheet4.cell(row=max_row4+1, column=2, value=input_field1.get())
    #customer name/particulars
    sheet4.cell(row=max_row4+1, column=3, value=input_field2.get())
    #debit linking
    formula_debit = f"='Cash Account Listing'!I{max_row3 + 1}"
    sheet4[f"D{max_row4 + 1}"] = formula_debit
    ######################################2023 turnover########################################################################
    #date
    sheet1.cell(row=max_row1+1, column=1, value=input_field.get())
    #invoice number
    sheet1.cell(row=max_row1+1, column=2, value=input_field1.get())
    #customer name/particulars
    sheet1.cell(row=max_row1+1, column=3, value=input_field2.get())
    #credit linking
    formula_credit = f"='Sales Invoice Listing'!BB{max_row + 1}"
    sheet1[f"E{max_row1 + 1}"] = formula_credit
    ######################################Debtors and Receivables########################################################################
    #date
    sheet2.cell(row=max_row2+1, column=1, value=input_field.get())
    #invoice number
    sheet2.cell(row=max_row2+1, column=2, value=input_field1.get())
    #customer name/particulars
    sheet2.cell(row=max_row2+1, column=3, value=input_field2.get())
    #debit
    sheet2.cell(row=max_row2+1, column=4, value=input_field55.get())
    #credit
    sheet2.cell(row=max_row2+1, column=5, value=input_field61.get())
    #balance
    formula_balance = f"=F{max_row2}-D{max_row2 + 1}+E{max_row2 + 1}"
    sheet2[f"F{max_row2 + 1}"] = formula_balance
    ####################################################################################################################################
 
    # Save the workbook
    workbook.save('LACOVA_MANAGEMENT_ACCOUNT_2023.xlsx')
    # messagebox.showinfo("Success", "Data pushed successfully!")
    global last_number
    last_number += 1
    with open("./images/delete.txt", "w") as file:
        file.write(str(last_number))
    # Update the submitted prompt
    submitted_label.config(text="Data pushed successfully")
    submitted_label.after(3000, lambda: submitted_label.config(text="")) # Remove the text after 3 seconds
    input_field.delete(0, tk.END)
    input_field1.delete(0, tk.END)
    input_field1.insert(tk.END, "{:07d}".format(last_number))
    input_field2.delete(0, tk.END)
    input_field3.delete(0, tk.END)
    input_field4.delete(0, tk.END)
    input_field5.delete(0, tk.END)
    input_field6.delete(0, tk.END)
    input_field7.delete(0, tk.END)
    input_field8.delete(0, tk.END)
    input_field9.delete(0, tk.END)
    input_field10.delete(0, tk.END)
    input_field11.delete(0, tk.END)
    input_field12.delete(0, tk.END)
    input_field13.delete(0, tk.END)
    input_field14.delete(0, tk.END)
    input_field15.delete(0, tk.END)
    input_field16.delete(0, tk.END)
    input_field17.delete(0, tk.END)
    input_field18.delete(0, tk.END)
    input_field19.delete(0, tk.END)
    input_field20.delete(0, tk.END)
    input_field21.delete(0, tk.END)
    input_field22.delete(0, tk.END)
    input_field23.delete(0, tk.END)
    input_field24.delete(0, tk.END)
    input_field25.delete(0, tk.END)
    input_field26.delete(0, tk.END)
    input_field27.delete(0, tk.END)
    input_field28.delete(0, tk.END)
    input_field29.delete(0, tk.END)
    input_field30.delete(0, tk.END)
    input_field31.delete(0, tk.END)
    input_field32.delete(0, tk.END)
    input_field33.delete(0, tk.END)
    input_field34.delete(0, tk.END)
    input_field35.delete(0, tk.END)
    input_field36.delete(0, tk.END)
    input_field37.delete(0, tk.END)
    input_field38.delete(0, tk.END)
    input_field39.delete(0, tk.END)
    input_field40.delete(0, tk.END)
    input_field41.delete(0, tk.END)
    input_field42.delete(0, tk.END)
    # input_field43.delete(0, tk.END)
    # input_field44.delete(0, tk.END)
    input_field45.delete(0, tk.END)
    input_field46.delete(0, tk.END)
    input_field47.delete(0, tk.END)
    input_field48.delete(0, tk.END)
    input_field49.delete(0, tk.END)
    input_field50.delete(0, tk.END)
    input_field51.delete(0, tk.END)
    input_field52.delete(0, tk.END)
    input_field53.delete(0, tk.END)
    input_field54.delete(0, tk.END)
    input_field55.delete(0, tk.END)
    input_field56.delete(0, tk.END)
    input_field57.delete(0, tk.END)
    input_field58.delete(0, tk.END)
    input_field59.delete(0, tk.END)
    input_field60.delete(0, tk.END)
    input_field61.delete(0, tk.END)
    input_field62.delete(0, tk.END)
    input_field63.delete(0, tk.END)
    input_field64.delete(0, tk.END)
    input_field65.delete(0, tk.END)
    messagebox.showinfo("Success", "Data pushed successfully!")
# Create the login window
login_window = tk.Tk()
login_window.geometry("475x420")
login_window.title("Lacova Login")
# Create the frame to hold all the widgets
frame = tk.Frame(login_window, bg="#3b5998")
frame.place(relx=0.5, rely=0.5, anchor="center")
# Create the logo label
logo_image = tk.PhotoImage(file="fb.png")
logo_label = tk.Label(frame, image=logo_image, bg="#3b5998")
logo_label.pack(pady=10)
# Create the labels and input fields for the username and password
username_label = tk.Label(frame, text="Username", fg="white", bg="#3b5998", font=("Helvetica", 12))
username_label.pack(pady=5)
username_entry = tk.Entry(frame, bg="#dfe3ee", font=("Helvetica", 12))
username_entry.pack(pady=5)
password_label = tk.Label(frame, text="Password", fg="white", bg="#3b5998", font=("Helvetica", 12))
password_label.pack(pady=5)
password_entry = tk.Entry(frame, show="*", bg="#dfe3ee", font=("Helvetica", 12))
password_entry.pack(pady=5)
submit_button = tk.Button(frame, text="Log In", command=check_credentials, bg="#1877f2", fg="white", font=("Helvetica", 12), width=25, bd=0)
submit_button.pack(pady=10)
# Create a label to display error messages
error_label = tk.Label(frame, fg="red")
error_label.pack(pady=5)
# Create a tkinter window and add input fields for the data
window = tk.Tk()
window.title("Lacova infinity Global Enterprises")
# Create the labels and input fields
tk.Label(window, text="Enter Date:").grid(row=0, column=0)
# input_field = DateEntry(window, width=12, background='darkblue', foreground='white', borderwidth=2)
input_field = tk.Entry(window)
input_field.grid(row=1, column=0)
tk.Label(window, text="Invoice Number:").grid(row=0, column=1)
input_field1 = tk.Entry(window)
input_field1.insert(tk.END, "{:07d}".format(last_number))
input_field1.grid(row=1, column=1)
tk.Label(window, text="Customer Name:").grid(row=0, column=2)
input_field2 = tk.Entry(window)
input_field2.grid(row=1, column=2)
tk.Label(window, text="QUANTITY of 3.4Kg:").grid(row=0, column=3)
input_field3 = tk.Entry(window)
input_field3.grid(row=1, column=3)
tk.Label(window, text="Price of 3.4Kg:").grid(row=0, column=4)
input_field4 = tk.Entry(window)
input_field4.grid(row=1, column=4)
tk.Label(window, text="QUANTITY of 3.5Kg:").grid(row=0, column=5)
input_field5 = tk.Entry(window)
input_field5.grid(row=1, column=5)
tk.Label(window, text="Price of 3.5Kg:").grid(row=2, column=0)
input_field6 = tk.Entry(window)
input_field6.grid(row=3, column=0)
tk.Label(window, text="QUANTITY of 3.6Kg:").grid(row=2, column=1)
input_field7 = tk.Entry(window)
input_field7.grid(row=3, column=1)
tk.Label(window, text="Price of 3.6Kg:").grid(row=2, column=2)
input_field8 = tk.Entry(window)
input_field8.grid(row=3, column=2)
tk.Label(window, text="QUANTITY of 3.7Kg:").grid(row=2, column=3)
input_field9 = tk.Entry(window)
input_field9.grid(row=3, column=3)
tk.Label(window, text="Price of 3.7Kg:").grid(row=2, column=4)
input_field10 = tk.Entry(window)
input_field10.grid(row=3, column=4)
tk.Label(window, text="QUANTITY of 3.8Kg:").grid(row=2, column=5)
input_field11 = tk.Entry(window)
input_field11.grid(row=3, column=5)
tk.Label(window, text="Price of 3.8Kg:").grid(row=4, column=0)
input_field12 = tk.Entry(window)
input_field12.grid(row=5, column=0)
tk.Label(window, text="QUANTITY of 3.9Kg:").grid(row=4, column=1)
input_field13 = tk.Entry(window)
input_field13.grid(row=5, column=1)
tk.Label(window, text="Price of 3.9Kg:").grid(row=4, column=2)
input_field14 = tk.Entry(window)
input_field14.grid(row=5, column=2)
tk.Label(window, text="QUANTITY of 4.0Kg:").grid(row=4, column=3)
input_field15 = tk.Entry(window)
input_field15.grid(row=5, column=3)
tk.Label(window, text="Price of 4.0Kg:").grid(row=4, column=4)
input_field16 = tk.Entry(window)
input_field16.grid(row=5, column=4)
tk.Label(window, text="QUANTITY of 4.1Kg:").grid(row=4, column=5)
input_field17 = tk.Entry(window)
input_field17.grid(row=5, column=5)
tk.Label(window, text="Price of 4.1Kg:").grid(row=6, column=0)
input_field18 = tk.Entry(window)
input_field18.grid(row=7, column=0)
tk.Label(window, text="QUANTITY of 4.2Kg:").grid(row=6, column=1)
input_field19 = tk.Entry(window)
input_field19.grid(row=7, column=1)
tk.Label(window, text="Price of 4.2Kg:").grid(row=6, column=2)
input_field20 = tk.Entry(window)
input_field20.grid(row=7, column=2)
tk.Label(window, text="QUANTITY of 4.3Kg:").grid(row=6, column=3)
input_field21 = tk.Entry(window)
input_field21.grid(row=7, column=3)
tk.Label(window, text="Price of 4.3Kg:").grid(row=6, column=4)
input_field22 = tk.Entry(window)
input_field22.grid(row=7, column=4)
tk.Label(window, text="QUANTITY of 4.4Kg:").grid(row=6, column=5)
input_field23 = tk.Entry(window)
input_field23.grid(row=7, column=5)
tk.Label(window, text="Price of 4.4Kg:").grid(row=8, column=0)
input_field24 = tk.Entry(window)
input_field24.grid(row=9, column=0)
tk.Label(window, text="QUANTITY of 4.5Kg:").grid(row=8, column=1)
input_field25 = tk.Entry(window)
input_field25.grid(row=9, column=1)
tk.Label(window, text="Price of 4.5Kg:").grid(row=8, column=2)
input_field26 = tk.Entry(window)
input_field26.grid(row=9, column=2)
tk.Label(window, text="QUANTITY of 4.6Kg:").grid(row=8, column=3)
input_field27 = tk.Entry(window)
input_field27.grid(row=9, column=3)
tk.Label(window, text="Price of 4.6Kg:").grid(row=8, column=4)
input_field28 = tk.Entry(window)
input_field28.grid(row=9, column=4)
tk.Label(window, text="QUANTITY of 4.7Kg:").grid(row=8, column=5)
input_field29 = tk.Entry(window)
input_field29.grid(row=9, column=5)
tk.Label(window, text="Price of 4.7Kg:").grid(row=10, column=0)
input_field30 = tk.Entry(window)
input_field30.grid(row=11, column=0)
tk.Label(window, text="QUANTITY of 4.8Kg:").grid(row=10, column=1)
input_field31 = tk.Entry(window)
input_field31.grid(row=11, column=1)
tk.Label(window, text="Price of 4.8Kg:").grid(row=10, column=2)
input_field32 = tk.Entry(window)
input_field32.grid(row=11, column=2)
tk.Label(window, text="QUANTITY of 4.9Kg:").grid(row=10, column=3)
input_field33 = tk.Entry(window)
input_field33.grid(row=11, column=3)
tk.Label(window, text="Price of 4.9Kg:").grid(row=10, column=4)
input_field34 = tk.Entry(window)
input_field34.grid(row=11, column=4)
tk.Label(window, text="QUANTITY of 5.0Kg:").grid(row=10, column=5)
input_field35 = tk.Entry(window)
input_field35.grid(row=11, column=5)
tk.Label(window, text="Price of 5.0Kg:").grid(row=12, column=0)
input_field36 = tk.Entry(window)
input_field36.grid(row=13, column=0)
tk.Label(window, text="QUANTITY of 5.2Kg:").grid(row=12, column=1)
input_field37 = tk.Entry(window)
input_field37.grid(row=13, column=1)
tk.Label(window, text="Price of 5.2Kg:").grid(row=12, column=2)
input_field38 = tk.Entry(window)
input_field38.grid(row=13, column=2)
tk.Label(window, text="QUANTITY of 5.5Kg:").grid(row=12, column=3)
input_field39 = tk.Entry(window)
input_field39.grid(row=13, column=3)
tk.Label(window, text="Price of 5.5Kg:").grid(row=12, column=4)
input_field40 = tk.Entry(window)
input_field40.grid(row=13, column=4)
tk.Label(window, text="QUANTITY of 5.8Kg:").grid(row=12, column=5)
input_field41 = tk.Entry(window)
input_field41.grid(row=13, column=5)
tk.Label(window, text="Price of 5.8Kg:").grid(row=14, column=0)
input_field42 = tk.Entry(window)
input_field42.grid(row=15, column=0)
tk.Label(window, text="QUANTITY of 6.0Kg:").grid(row=14, column=1)
input_field45 = tk.Entry(window)
input_field45.grid(row=15, column=1)
tk.Label(window, text="Price of 6.0Kg:").grid(row=14, column=2)
input_field46 = tk.Entry(window)
input_field46.grid(row=15, column=2)
tk.Label(window, text="QUANTITY of Oloja:").grid(row=14, column=3)
input_field47 = tk.Entry(window)
input_field47.grid(row=15, column=3)
tk.Label(window, text="Price of Oloja:").grid(row=14, column=4)
input_field48 = tk.Entry(window)
input_field48.grid(row=15, column=4)
tk.Label(window, text="QUANTITY of Ice:").grid(row=14, column=5)
input_field49 = tk.Entry(window)
input_field49.grid(row=15, column=5)
tk.Label(window, text="Price of Ice:").grid(row=16, column=0)
input_field50 = tk.Entry(window)
input_field50.grid(row=17, column=0)
tk.Label(window, text="QUANTITY of Dispenser:").grid(row=16, column=1)
input_field51 = tk.Entry(window)
input_field51.grid(row=17, column=1)
tk.Label(window, text="Price of Dispenser:").grid(row=16, column=2)
input_field52 = tk.Entry(window)
input_field52.grid(row=17, column=2)
tk.Label(window, text="QUANTITY of Small:").grid(row=16, column=3)
input_field53 = tk.Entry(window)
input_field53.grid(row=17, column=3)
tk.Label(window, text="Price of Small:").grid(row=16, column=4)
input_field54 = tk.Entry(window)
input_field54.grid(row=17, column=4)
# tk.Label(window, text="TOTAL").grid(row=16, column=5)
# Create the button to calculate the total
calculate_button = tk.Button(window, text="Cal. Total Price", command=calculate_total, bg="blue", fg="white", font=("Helvetica", 12), width=10, bd=0)
calculate_button.grid(row=16, column=5)
input_field55 = tk.Entry(window)
input_field55.grid(row=17, column=5)
tk.Label(window, text="FCMB").grid(row=18, column=0)
input_field56 = tk.Entry(window)
input_field56.grid(row=19, column=0)
tk.Label(window, text="Polaris Bank").grid(row=18, column=1)
input_field57 = tk.Entry(window)
input_field57.grid(row=19, column=1)
tk.Label(window, text="Access Bank").grid(row=18, column=2)
input_field58 = tk.Entry(window)
input_field58.grid(row=19, column=2)
tk.Label(window, text="Amount Paid in Cash").grid(row=18, column=3)
input_field59 = tk.Entry(window)
input_field59.grid(row=19, column=3)
tk.Label(window, text="Details").grid(row=18, column=4)
input_field60 = tk.Entry(window)
input_field60.grid(row=19, column=4)
# Create the button to calculate the total_invoice
calculate_button = tk.Button(window, text="Cal. Total Amount Paid", command=calculate_total_invoice_amount, bg="blue", fg="white", font=("Helvetica", 12), width=10, bd=0)
calculate_button.grid(row=18, column=5)
input_field61 = tk.Entry(window)
input_field61.grid(row=19, column=5)
#############old balance#####################
tk.Label(window, text="Old Balance").grid(row=20, column=0)
input_field62 = tk.Entry(window)
input_field62.grid(row=21, column=0)
# Create the button to calculate the balances
balances_button = tk.Button(window, text="Calculate Balances", command=calculate_balances, bg="blue", fg="white", font=("Helvetica", 12), width=10, bd=0)
balances_button.grid(row=21, column=1)
#############outstanding#####################
tk.Label(window, text="Outstanding").grid(row=20, column=2)
input_field63 = tk.Entry(window)
input_field63.grid(row=21, column=2)
#############prepayment#####################
tk.Label(window, text="Prepayment").grid(row=20, column=3)
input_field64 = tk.Entry(window)
input_field64.grid(row=21, column=3)
#############New Balance#####################
tk.Label(window, text="New Balance").grid(row=20, column=4)
input_field65 = tk.Entry(window)
input_field65.grid(row=21, column=4)
submitted_label = tk.Label(window, text="")
submitted_label.grid(row=22, column=1, columnspan=3)
not_submitted_label = tk.Label(window, text="")
not_submitted_label.grid(row=22, column=2, columnspan=3)
# Add a button to log out
logout_button = tk.Button(window, text="Log Out", command=logout, bg="red", fg="white", font=("Helvetica", 12), width=10, bd=0)
logout_button.grid(row=24, column=5, columnspan=3)
# Create a button widget to select the file and save the data
select_button = tk.Button(window, text="Push to Database", command=save_data, bg="green", fg="white", font=("Helvetica", 12), width=10, bd=0)
select_button.grid(row=24, column=3)
# Create a button to generate the receipt
generate_button = tk.Button(window, text="Generate Receipt", command=generate_receipt, bg="green", fg="white", font=("Helvetica", 12), width=10, bd=0)
generate_button.grid(row=24, column=0)
# Create a button to generate the Waybill
generate_button = tk.Button(window, text="Generate Waybill", command=generate_Waybill, bg="green", fg="white", font=("Helvetica", 12), width=10, bd=0)
generate_button.grid(row=24, column=1)


for child in window.winfo_children():
    child.grid_configure(padx=5, pady=5, sticky="nsew")
    window.grid_columnconfigure(child.grid_info()["column"], weight=1)
    window.grid_rowconfigure(child.grid_info()["row"], weight=1)
login_window.grid_anchor("center")
# Run the tkinter mainloop
# window.mainloop()
window.withdraw()
# Start the login window
# login_window.mainloop()
window.mainloop()

